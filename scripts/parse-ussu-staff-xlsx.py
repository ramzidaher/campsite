#!/usr/bin/env python3
"""Parse USSU-style staff workbook (one sheet named "Staff") and emit JSON for provision-university-of-sussex-student-union.mjs."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl. Install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def norm(s: object | None) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split())


PAY_ONLY = frozenset({"Part-time", "Weekly Paid", "Weekly Paid/PT"})


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Staff" not in wb.sheetnames:
        raise SystemExit(f'Expected a "Staff" sheet; found {wb.sheetnames!r}')
    ws = wb["Staff"]
    rows = list(ws.iter_rows(values_only=True))
    structural_dept: str | None = None
    out: list[dict] = []

    for i, r in enumerate(rows):
        if i < 2:
            continue
        if r[0] and norm(r[0]):
            sec = norm(r[0])
            if sec not in PAY_ONLY:
                structural_dept = sec
        if not r[2] or not norm(r[2]):
            continue
        name = norm(r[2])
        job = norm(r[3]) if r[3] else ""
        if not job:
            continue
        if re.match(r"^vacant", name, re.I):
            continue

        def iso(x: object) -> str | None:
            if hasattr(x, "isoformat"):
                return x.isoformat()  # type: ignore[no-any-return]
            return None

        row = {
            "sheet_row": i + 1,
            "structural_dept": structural_dept,
            "profile_note": norm(r[1]) if r[1] else None,
            "full_name": name,
            "job_title": job,
            "contract_type_raw": norm(r[4]) if r[4] else None,
            "end_date": iso(r[5]) if r[5] else None,
            "grade": str(r[6]).strip() if r[6] is not None else None,
            "point": str(r[7]).strip() if r[7] is not None else None,
            "manager_text": norm(r[10]) if r[10] else None,
            "budget_code": norm(r[14]) if r[14] else None,
            "hours_text": norm(r[15]) if r[15] else None,
            "employed_by": norm(r[16]) if r[16] else None,
            "role_start": iso(r[17]) if r[17] else None,
            "continuous_employment": iso(r[18]) if r[18] else None,
        }
        out.append(row)
    return out


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: parse-ussu-staff-xlsx.py <path.xlsx>", file=sys.stderr)
        sys.exit(1)
    data = parse_workbook(sys.argv[1])
    json.dump(data, sys.stdout, indent=2)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
