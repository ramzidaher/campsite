#!/usr/bin/env python3
"""
Build a CSV in Google Chrome “Passwords.csv” export format (import into Chrome / Google Password Manager).

Header matches Chrome:
  name,url,username,password,note

Same columns as repo “Chrome Passwords.csv” (see first line of that file).

Emails match provision-university-of-sussex-student-union.mjs when the workbook and password are the same.

Usage:
  CAMPSITE_USSU_PASSWORD='…' python3 scripts/generate-ussu-password-import-tsv.py
  python3 scripts/generate-ussu-password-import-tsv.py [path/to/staff.xlsx] [password]
"""

from __future__ import annotations

import csv
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl. Install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

_ROOT = Path(__file__).resolve().parent.parent

PAY_ONLY = frozenset({"Part-time", "Weekly Paid", "Weekly Paid/PT"})


def norm(s: object | None) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split())


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Staff" not in wb.sheetnames:
        raise SystemExit(f'Expected a "Staff" sheet; found {wb.sheetnames!r}')
    ws = wb["Staff"]
    rows = list(ws.iter_rows(values_only=True))
    structural_dept: str | None = None
    out: list[dict] = []

    for i, r in enumerate(rows):
        if i < 2:
            continue
        if r[0] and norm(r[0]):
            sec = norm(r[0])
            if sec not in PAY_ONLY:
                structural_dept = sec
        if not r[2] or not norm(r[2]):
            continue
        name = norm(r[2])
        job = norm(r[3]) if r[3] else ""
        if not job:
            continue
        if re.match(r"^vacant", name, re.I):
            continue
        out.append({"sheet_row": i + 1, "structural_dept": structural_dept, "full_name": name})
    return out


def allocate_email(full_name: str, used_local: dict[str, int]) -> str:
    parts = [p for p in full_name.strip().lower().split() if p]
    raw_first = re.sub(r"[^a-z\-]", "", parts[0] if parts else "") or "user"
    raw_last = re.sub(r"[^a-z\-]", "", parts[-1] if parts else "") or "user"
    base = re.sub(r"^\.|\.$", "", f"{raw_first}.{raw_last}") or "user.user"
    n = used_local.get(base, 0)
    used_local[base] = n + 1
    suffix = "" if n == 0 else str(n + 1)
    return f"{base}{suffix}@camp-site.co.uk"


def default_xlsx() -> Path:
    name = "HR Copy of Staff List (excluding trading CSA's & DM's) copy.xlsx"
    return _ROOT / name


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx()
    password = (
        sys.argv[2]
        if len(sys.argv) > 2
        else os.environ.get("CAMPSITE_USSU_PASSWORD", "").strip() or "Campsite@2026!"
    )
    if not xlsx.is_file():
        print(f"Workbook not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    rows = parse_workbook(str(xlsx))
    used: dict[str, int] = {}
    rows_sorted = sorted(rows, key=lambda r: int(r.get("sheet_row") or 0))

    site_name = "camp-site.co.uk"
    login_url = "https://camp-site.co.uk/login"
    note = (
        "University of Sussex Student Union (USSU) seed; shared password for all rows "
        "unless you changed CAMPSITE_USSU_PASSWORD when seeding."
    )

    out_dir = Path(__file__).resolve().parent / "ussu-provision-output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "ussu-password-import.csv"

    n = 0
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "url", "username", "password", "note"])
        for r in rows_sorted:
            full_name = str(r["full_name"])
            email = allocate_email(full_name, used)
            w.writerow([site_name, login_url, email, password, note])
            n += 1

    print(f"Wrote {n} data rows (+ header) to {out_path}")


if __name__ == "__main__":
    main()
